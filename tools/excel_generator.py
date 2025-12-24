"""
📁 Excel Generator Tool
Genera archivos Excel profesionales con pandas y openpyxl
"""
from typing import Dict, Any, List, Optional
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Fill, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from pathlib import Path
from datetime import datetime
import base64
import config
from utils.logger import get_logger

logger = get_logger("ExcelGenerator")


class ExcelGenerator:
    """Generador de archivos Excel profesionales"""
    
    def __init__(self):
        """Inicializa el generador"""
        self.output_dir = Path(config.EXPORTS_DIR)
        self.output_dir.mkdir(exist_ok=True)
    
    def create_excel_from_data(
        self,
        data: pd.DataFrame,
        filename: Optional[str] = None,
        sheet_name: str = None,
        include_index: bool = False,
        auto_filter: bool = True,
        freeze_header: bool = True,
        apply_styling: bool = True
    ) -> Dict[str, Any]:
        """
        Crea archivo Excel desde DataFrame
        
        Args:
            data: DataFrame con los datos
            filename: Nombre del archivo (sin extensión)
            sheet_name: Nombre de la hoja
            include_index: Si True, incluye el índice
            auto_filter: Si True, agrega autofiltro
            freeze_header: Si True, congela primera fila
            apply_styling: Si True, aplica estilos profesionales
            
        Returns:
            Dict con información del archivo creado
        """
        try:
            # Generar nombre de archivo
            if not filename:
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                filename = f"export_{timestamp}"
            
            file_path = self.output_dir / f"{filename}.xlsx"
            sheet_name = sheet_name or config.EXCEL_DEFAULT_SHEET_NAME
            
            # Crear archivo Excel
            with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                data.to_excel(
                    writer,
                    sheet_name=sheet_name,
                    index=include_index,
                    freeze_panes=(1, 0) if freeze_header else None
                )
            
            # Aplicar formato si se solicita
            if apply_styling:
                self._apply_professional_styling(
                    file_path,
                    sheet_name,
                    auto_filter=auto_filter,
                    freeze_header=freeze_header
                )
            
            logger.info(f"✅ Excel creado: {filename}.xlsx con {len(data)} filas")
            
            return {
                "success": True,
                "file_path": str(file_path),
                "filename": f"{filename}.xlsx",
                "row_count": len(data),
                "column_count": len(data.columns),
                "sheet_count": 1,
                "sheets": [sheet_name]
            }
        
        except Exception as e:
            logger.error(f"❌ Error al crear Excel: {str(e)}")
            return {
                "success": False,
                "error": str(e)
            }
    
    def create_multi_sheet_excel(
        self,
        sheets_data: Dict[str, pd.DataFrame],
        filename: Optional[str] = None,
        apply_styling: bool = True
    ) -> Dict[str, Any]:
        """
        Crea archivo Excel con múltiples hojas
        
        Args:
            sheets_data: Dict donde key=nombre_hoja, value=DataFrame
            filename: Nombre del archivo (sin extensión)
            apply_styling: Si True, aplica estilos
            
        Returns:
            Dict con información del archivo
        """
        try:
            # Generar nombre de archivo
            if not filename:
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                filename = f"multi_sheet_{timestamp}"
            
            file_path = self.output_dir / f"{filename}.xlsx"
            
            # Crear archivo Excel con múltiples hojas
            with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                for sheet_name, df in sheets_data.items():
                    df.to_excel(
                        writer,
                        sheet_name=sheet_name[:31],  # Excel limit
                        index=False
                    )
            
            # Aplicar formato a cada hoja
            if apply_styling:
                for sheet_name in sheets_data.keys():
                    self._apply_professional_styling(
                        file_path,
                        sheet_name[:31],
                        auto_filter=True,
                        freeze_header=True
                    )
            
            total_rows = sum(len(df) for df in sheets_data.values())
            
            logger.info(
                f"✅ Excel multi-hoja creado: {filename}.xlsx con {len(sheets_data)} hojas"
            )
            
            return {
                "success": True,
                "file_path": str(file_path),
                "filename": f"{filename}.xlsx",
                "row_count": total_rows,
                "sheet_count": len(sheets_data),
                "sheets": list(sheets_data.keys())
            }
        
        except Exception as e:
            logger.error(f"❌ Error al crear Excel multi-hoja: {str(e)}")
            return {
                "success": False,
                "error": str(e)
            }
    
    def _apply_professional_styling(
        self,
        file_path: Path,
        sheet_name: str,
        auto_filter: bool = True,
        freeze_header: bool = True
    ):
        """
        Aplica estilos profesionales al Excel
        
        Args:
            file_path: Path del archivo
            sheet_name: Nombre de la hoja
            auto_filter: Si True, aplica autofiltro
            freeze_header: Si True, congela encabezado
        """
        try:
            wb = load_workbook(file_path)
            ws = wb[sheet_name]
            
            # Color corporativo
            header_fill = PatternFill(
                start_color="1F77B4",
                end_color="1F77B4",
                fill_type="solid"
            )
            
            # Estilos de encabezado
            header_font = Font(
                name='Arial',
                size=11,
                bold=True,
                color="FFFFFF"
            )
            
            header_alignment = Alignment(
                horizontal='center',
                vertical='center',
                wrap_text=True
            )
            
            # Bordes
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Aplicar estilo al encabezado (primera fila)
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment
                cell.border = thin_border
            
            # Ajustar ancho de columnas
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                adjusted_width = min(max_length + 2, 50)  # Max 50
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Aplicar bordes a todas las celdas con datos
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
                for cell in row:
                    cell.border = thin_border
            
            # Autofiltro
            if auto_filter and ws.max_row > 1:
                ws.auto_filter.ref = ws.dimensions
            
            # Congelar encabezado
            if freeze_header:
                ws.freeze_panes = 'A2'
            
            # Agregar metadatos
            wb.properties.creator = config.EXCEL_AUTHOR
            wb.properties.company = config.EXCEL_COMPANY
            wb.properties.created = datetime.now()
            
            wb.save(file_path)
            logger.info(f"✅ Estilos aplicados a {sheet_name}")
        
        except Exception as e:
            logger.error(f"❌ Error al aplicar estilos: {str(e)}")
    
    def add_chart_to_excel(
        self,
        file_path: str,
        sheet_name: str,
        chart_type: str,
        data_range: str,
        chart_title: str,
        position: str = "E2"
    ) -> Dict[str, Any]:
        """
        Agrega un gráfico a una hoja de Excel existente
        
        Args:
            file_path: Path del archivo Excel
            sheet_name: Nombre de la hoja
            chart_type: Tipo de gráfico ('bar', 'line', 'pie')
            data_range: Rango de datos (ej: 'A1:B10')
            chart_title: Título del gráfico
            position: Posición donde insertar (ej: 'E2')
            
        Returns:
            Dict con resultado
        """
        try:
            wb = load_workbook(file_path)
            ws = wb[sheet_name]
            
            # Crear gráfico según tipo
            if chart_type == "bar":
                chart = BarChart()
            elif chart_type == "line":
                chart = LineChart()
            elif chart_type == "pie":
                chart = PieChart()
            else:
                return {"success": False, "error": f"Tipo de gráfico '{chart_type}' no soportado"}
            
            chart.title = chart_title
            chart.style = 10
            
            # Definir datos y categorías
            # Esto es un ejemplo simple, puede necesitar ajustes según la estructura
            data = Reference(ws, min_col=2, min_row=1, max_row=ws.max_row)
            categories = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)
            
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(categories)
            
            # Insertar gráfico
            ws.add_chart(chart, position)
            
            wb.save(file_path)
            logger.info(f"✅ Gráfico {chart_type} agregado a {sheet_name}")
            
            return {
                "success": True,
                "chart_type": chart_type,
                "sheet": sheet_name
            }
        
        except Exception as e:
            logger.error(f"❌ Error al agregar gráfico: {str(e)}")
            return {
                "success": False,
                "error": str(e)
            }
    
    def export_as_base64(self, file_path: str) -> Optional[str]:
        """
        Convierte archivo Excel a base64 para envío
        
        Args:
            file_path: Path del archivo
            
        Returns:
            String base64 o None si hay error
        """
        try:
            with open(file_path, 'rb') as f:
                file_data = f.read()
            
            base64_str = base64.b64encode(file_data).decode('utf-8')
            logger.info(f"✅ Excel convertido a base64")
            return base64_str
        
        except Exception as e:
            logger.error(f"❌ Error al convertir a base64: {str(e)}")
            return None
    
    def create_summary_sheet(
        self,
        data: pd.DataFrame,
        summary_stats: Dict[str, Any]
    ) -> pd.DataFrame:
        """
        Crea hoja de resumen con estadísticas
        
        Args:
            data: DataFrame original
            summary_stats: Dict con estadísticas personalizadas
            
        Returns:
            DataFrame con resumen
        """
        summary_data = {
            "Métrica": [],
            "Valor": []
        }
        
        # Estadísticas básicas
        summary_data["Métrica"].append("Total de Registros")
        summary_data["Valor"].append(len(data))
        
        summary_data["Métrica"].append("Total de Columnas")
        summary_data["Valor"].append(len(data.columns))
        
        # Agregar estadísticas personalizadas
        for key, value in summary_stats.items():
            summary_data["Métrica"].append(key)
            summary_data["Valor"].append(value)
        
        return pd.DataFrame(summary_data)


# Singleton instance
_excel_generator_instance = None

def get_excel_generator() -> ExcelGenerator:
    """
    Obtiene instancia del generador (singleton)
    
    Returns:
        ExcelGenerator instance
    """
    global _excel_generator_instance
    if _excel_generator_instance is None:
        _excel_generator_instance = ExcelGenerator()
    return _excel_generator_instance
